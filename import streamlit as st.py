import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import pytesseract
import io
import pandas as pd

# Configuración de Tesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\\tesseract.exe"

archivo_excel = "libros.xlsx"

# Usuarios con roles
USUARIOS = {
    "admin": {"password": "1234", "rol": "admin"},
    "maikel": {"password": "abcd", "rol": "admin"},
    "lector": {"password": "0000", "rol": "lector"}
}

# Crear archivo si no existe
def crear_archivo():
    try:
        wb = openpyxl.load_workbook(archivo_excel)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Libros"
        ws.append(["Título", "Autor", "Editorial", "Foto Portada"])
        wb.save(archivo_excel)

crear_archivo()

def listar_libros():
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    data = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        data.append(fila)
    df = pd.DataFrame(data, columns=["Título", "Autor", "Editorial", "Foto Portada"])
    return df

def registrar_libro(titulo, autor, editorial, foto):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    fila = ws.max_row + 1
    ws.cell(row=fila, column=1, value=titulo)
    ws.cell(row=fila, column=2, value=autor)
    ws.cell(row=fila, column=3, value=editorial)

    img_bytes = io.BytesIO(foto.read())
    img_temp = PILImage.open(img_bytes)
    img_temp.save("temp_portada.png")
    img_excel = ExcelImage("temp_portada.png")
    img_excel.width, img_excel.height = 100, 150
    ws.add_image(img_excel, f"D{fila}")

    wb.save(archivo_excel)

def actualizar_portada(titulo, foto):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=1).value == titulo:
            img_bytes = io.BytesIO(foto.read())
            img_temp = PILImage.open(img_bytes)
            img_temp.save("temp_portada.png")
            img_excel = ExcelImage("temp_portada.png")
            img_excel.width, img_excel.height = 100, 150
            ws.add_image(img_excel, f"D{fila}")
            wb.save(archivo_excel)
            return True
    return False

def eliminar_libro(titulo):
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb["Libros"]
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=1).value == titulo:
            ws.delete_rows(fila)
            wb.save(archivo_excel)
            return True
    return False

# ------------------- LOGIN -------------------
st.title("📚 Registro de Libros con OCR")

if "logueado" not in st.session_state:
    st.session_state.logueado = False
    st.session_state.rol = None

if not st.session_state.logueado:
    st.subheader("🔐 Iniciar sesión")
    usuario = st.text_input("Usuario")
    clave = st.text_input("Contraseña", type="password")
    if st.button("Login"):
        if usuario in USUARIOS and USUARIOS[usuario]["password"] == clave:
            st.session_state.logueado = True
            st.session_state.rol = USUARIOS[usuario]["rol"]
            st.success(f"✅ Acceso concedido como {st.session_state.rol}")
        else:
            st.error("❌ Usuario o contraseña incorrectos")
else:
    rol = st.session_state.rol
    st.info(f"👤 Rol actual: {rol}")

    # ------------------- APP PRINCIPAL -------------------
    if rol == "admin":
        foto = st.file_uploader("Sube la foto de la portada", type=["jpg","jpeg","png"])
        if foto:
            st.image(foto, caption="Portada subida", width="stretch")
            imagen = PILImage.open(foto)
            texto = pytesseract.image_to_string(imagen, lang="spa")
            st.text_area("Texto detectado", texto)

            lineas = [l.strip() for l in texto.split("\n") if l.strip()]
            titulo = st.text_input("Título", lineas[0] if len(lineas) > 0 else "")
            autor = st.text_input("Autor", lineas[1] if len(lineas) > 1 else "")
            editorial = st.text_input("Editorial", lineas[2] if len(lineas) > 2 else "")

            if st.button("Guardar Registro"):
                registrar_libro(titulo, autor, editorial, foto)
                st.success("✅ Libro registrado correctamente con imagen en Excel.")

    # Listado con botones según rol
    st.subheader("📊 Listado de Libros Registrados")
    df_libros = listar_libros()
    for i, row in df_libros.iterrows():
        col1, col2, col3, col4, col5 = st.columns([3, 3, 3, 1, 2])
        col1.write(row["Título"])
        col2.write(row["Autor"])
        col3.write(row["Editorial"])

        if rol == "admin":
            if col4.button("🗑️ Eliminar", key=f"del_{i}"):
                if eliminar_libro(row["Título"]):
                    st.success(f"Libro '{row['Título']}' eliminado correctamente.")
                    st.experimental_rerun()

            nueva_portada = col5.file_uploader(f"Configurar portada {row['Título']}", type=["jpg","jpeg","png"], key=f"portada_{i}")
            if nueva_portada and col5.button("Actualizar", key=f"upd_{i}"):
                if actualizar_portada(row["Título"], nueva_portada):
                    st.success(f"✅ Portada de '{row['Título']}' actualizada.")
                    st.experimental_rerun()
